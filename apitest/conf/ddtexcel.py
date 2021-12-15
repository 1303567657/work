from openpyxl import load_workbook


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        #print(sheet.max_row)
        #print(sheet.cell( 1, 2).value)
        test_data = []
        for i in range(1, sheet.max_row):
            sub_data = {}
            sub_data['method'] = sheet.cell(i + 1, 2).value

            sub_data['url'] = sheet.cell(i + 1, 3).value
            sub_data['data'] = sheet.cell(i + 1, 4).value
            sub_data['ass'] = sheet.cell(i + 1, 5).value
            # sub_data['purpose'] = sheet.cell(i + 1, 6).value  # 验证目的
            test_data.append(sub_data)
        return test_data
test_data = DoExcel(r'D:\work\study\pycharm\seleniumstudy\apitest\data\case.xlsx', "one").get_data()#获取到Excel表里的全部数据
print(test_data)
# import xlrd
#
#
# def get_data(file_name):
#     aa = []
#     book = xlrd.open_workbook(file_name)
#     sheet = book.sheet_by_index(0)
#     for i in range(1, sheet.nrows):
#         aa.append(list(sheet.row_values(i, 0, sheet.ncols)))
#     return aa
#
#
# result = get_data(r'D:\work\study\pycharm\seleniumstudy\apitest\data\case.xlsx')
# print(result)
